import os
import ast
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


from doctors_result import doctor_result_dict


    #вместо этой функции необходимо написать работающий код
def ctg_analyze(df_all_coords):
    
    # Исправление датафрейма
    df_all_coords = df_all_coords[df_all_coords['y'] >  0]

    # Число фишера
    fisher_score = 0
    
    # Базальный ритм

    baseline_HR = df_all_coords['y'].mean()

    # Подсчёт
    if baseline_HR >= 120 and baseline_HR <= 160:
        fisher_score += 2
    elif baseline_HR >= 100 and baseline_HR <= 180 :
        fisher_score += 1
    
    
    # Амплитуда осцилляций
    df_amp = df_all_coords['y'] - baseline_HR
    amp = df_amp.abs().max()

    # Подсчёт
    if amp >= 10 and amp <= 25:
        fisher_score += 2
    elif (amp >= 5 and amp <= 9) or amp > 25:
        fisher_score += 1

    # Частота осцилляций
    df_rolled = df_all_coords['y'].rolling(10, center=True).mean().round().dropna()
    last = 0
    osc_flag = 0
    osc_count = 0
    for i in df_rolled.index:
        if df_rolled[i]  != last:
            if df_rolled[i] > last and osc_flag != 1:
                osc_flag = 1
                last = df_rolled[i] 
                osc_count += 1
            elif df_rolled[i]  < last and osc_flag != -1:
                osc_flag = -1
                last = df_rolled[i] 
                osc_count += 1
            else:
                last = df_rolled[i] 
                
    osc_count = osc_count / (df_all_coords['x'].max() / 60)

    # Подсчёт
    if osc_count >= 6:
        fisher_score += 2
    elif osc_count >= 3 and osc_count < 6:
        fisher_score += 1

    
    # Акцелерации
    f = 0
    acc_count = 0
    df_all_coords.insert(loc=len(df_all_coords.columns), column='amp', value=df_amp)
    
    for acci in df_all_coords[df_all_coords['x'] < 1200].index:
        if f == 0 and df_all_coords.iloc[acci,2] >= 15:
            timestart = df_all_coords.iloc[acci,0]
            f = 1
        if f == 1 and df_all_coords.iloc[acci,2] < 15:
            f = 0
            if df_all_coords.iloc[acci,0] - timestart > 15:
                acc_count += 1
    acc_count = acc_count / 2

    # Подсчёт
    if acc_count >= 2:
        fisher_score += 2
    elif osc_count < 2:
        fisher_score += 1

    # Децелерации
    f = 0
    dec_count = 0
    for deci in df_all_coords[df_all_coords['x'] < 1200].index:
        if f == 0 and df_all_coords.iloc[deci,2] <= -15:
            timestart = df_all_coords.iloc[deci,2]
            f = 1
        if f == 1 and df_all_coords.iloc[deci,2] > -15:
            f = 0
            if df_all_coords.iloc[deci,2] - timestart > 15:
                dec_count += 1
    dec_count = dec_count / 2

        # Подсчёт
    if dec_count < 1:
        fisher_score += 2
    elif dec_count <= 2:
        fisher_score += 1

    

    if fisher_score >= 8:
        return 'хорошее'
    else:
        return 'плохое'



if __name__ == '__main__':

    directory = 'ctg_files'
    program_result_dict = {}
    start_time = datetime.now()
    #проходим циклом по предоставленным файлам с массивами данных по КТГ
    filename_list = os.listdir(directory)
    filename_list.sort(key=lambda x: int(x[:-4]))
    for filename in filename_list:
        f = os.path.join(directory, filename)
        if os.path.isfile(f):
            file = open(f, 'r')
            graph_list = ast.literal_eval(file.read())
            #преобразуем данные в pandas dataframe для дальнейшей обработки
            #преобразовывать в dataframe необязательно, если имеются другие решения можете реализовать их
            x_coords = [i.get('Key') for i in graph_list]
            y_coords = [i.get('Value') for i in graph_list]
            df_all_coords = pd.DataFrame.from_dict({'x': x_coords, 
                                                    'y': y_coords, 
                                                    })


            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////
            '''здесь вызывается исполнение функции оценивающей КТГ
            при написании кода рекомендуется использование matplotlib или аналоги для визуализации графика, это поможет
            писать весь код в одной функции необязательно - хорошая читаемость кода приветствуется
            программа в результате должна вернуть строку 'хорошее' или 'плохое'
            '''
            program_result = ctg_analyze(df_all_coords)
            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////


            #записывается результирующий словарь с ключами идентичными словарю doctor_result_dict для дальнейшего сравнения
            program_result_dict[filename] = program_result

    #считаем среднее время выполнения оценки одного КТГ
    average_time = (datetime.now() - start_time) / len(os.listdir(directory))
    print(f'среднее время выполнения оценки одного КТГ - {average_time}')

    #считаем количество совпадений программы с врачом
    number_of_matches = 0
    for res in program_result_dict:
        if program_result_dict[res] == doctor_result_dict[res]:
            number_of_matches += 1
    print(f'совпадений программы с врачом {number_of_matches} из 100')

    #в таблицу сохраняется результат
    #в ней можно будет более подробно рассмотреть общую картину того в каких случаях расхождения между врачом и программой
    wb = load_workbook('ctg.xlsx')
    del wb['Sheet1']
    ws = wb.create_sheet('Sheet1')
    for res in program_result_dict:
        ws.append([res, doctor_result_dict[res], program_result_dict[res]])
    wb.save('ctg.xlsx')
